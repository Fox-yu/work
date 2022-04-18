import os
import time
import pymysql
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font

def searchFileContent(filename):
    try:
        fp1 = open(filename, 'a', encoding="utf-8")
        fp1.write('\n')
        fp1.close()
        fp =open(filename, 'r', encoding="utf-8")
        fr = fp.read()
        confHead = re.findall('#.*?\n+\[', fr)
        confCenter = re.split(r'#.*?\n\[', fr)  # 分割配置
        confCenter.remove(confCenter[0])  # 移除空数据
        confIdTemp = []
        createDate = time.strftime("%Y%m%d", time.localtime())  # 创建日期
        wb = Workbook()
        ws = wb.active
        ws.title = filename
        title = ['厂商名称', '父协议编号', '应用类别', '协议编号', '应用', '应用曾用名', '终端', '版本号', '动作值', '动作描述\n（部分终端版本支持）', '触发应用',
                 '应用触发动作',
                 '还原条件', '需求来源', '协议', '实现方式', '分析日期', '分析负责人', '分析方式', '配置ID', '开发日期', '开发负责人', '开发周期', '测试日期',
                 '测试负责人',
                 '还原系统版本号', '还原功能类型', '数据来源', '状态', '配置/代码变动', '逆向状态', '逆向负责人', '逆向日期', '逆向周期', 'BUG修复人', 'BUG修复日期',
                 'BUG修复周期(h)', '未实现原因', '实现文件', '加密类型', '失效版本号', '全文协议值', '全文类型', '全文后缀', '附件类别', '坐标系', '是否纠偏',
                 '关闭应用权限是否上报硬件信息', '关闭应用权限是否上报经纬度信息', '用户层面COOKIE来源', '会话层面COOKIE来源', '备注', '还原字段', '字段名']  # 表头
        font = Font("黑体", size=14, bold=True)
        font2 = Font("等线", size=11)
        for col in range(len(title)):
            c = col + 1
            ws.cell(row=1, column=c).value = title[col]
            ws.cell(row=1, column=c).font = font  # 插入表头
        i = 1
        # 数据切片
        confDomainCount = 0
        fieldCount = []
        hostCount = 0
        urlCount = 0
        for (line, line2) in zip(confCenter, confHead):
            # 连接数据库
            conn_field = pymysql.connect(host="127.0.0.1", port=3306, user="root", passwd="980097", db="field",
                                         charset="utf8")
            cursor_field = conn_field.cursor(pymysql.cursors.DictCursor)
            conn_action = pymysql.connect(host="127.0.0.1", port=3306, user="root", passwd="980097", db="action",
                                          charset="utf8")
            cursor_action = conn_action.cursor(pymysql.cursors.DictCursor)
            conn_protocol = pymysql.connect(host="127.0.0.1", port=3306, user="root", passwd="980097", db="protocol",
                                            charset="utf8")
            cursor_protocol = conn_protocol.cursor(pymysql.cursors.DictCursor)
            conn_user = pymysql.connect(host="127.0.0.1", port=3306, user="root", passwd="980097", db="user",
                                        charset="utf8")
            cursor_user = conn_user.cursor(pymysql.cursors.DictCursor)

            i = i + 1
            # 配置整体
            conf = line2 + line
            # 配置域分割
            try:
                confTitle = re.findall('\[(.*?)\]', conf)[0].split('-')  # 配置域
                toolType = confTitle[0]  # 应用终端
                appVersion = confTitle[1]  # 应用版本
                date = time.strftime("%Y/%m/%d", time.localtime())  # 日期
                if toolType == 'iosweb':
                    toolType = 'IOS网页'
                elif toolType == 'androidweb':
                    toolType = 'ANDROID网页'
                elif toolType == 'pcweb':
                    toolType = 'PC网页'
                elif toolType == 'ios':
                    toolType = 'IOS客户端'
                elif toolType == 'android':
                    toolType = 'ANDROID客户端'
                elif toolType == 'pc':
                    toolType = 'PC客户端'
                elif toolType == 'android&ios':
                    toolType = 'ANDROID客户端'
                elif toolType == 'ios&android':
                    toolType = 'IOS客户端'
                else:
                    print(line2 + '终端命名错误，请修改配置]')
                    break
            except IndexError:
                print(line2 + "当前配置域错误，请修改配置]")  # 格式化终端
            try:
                actionTrigger = re.findall('#(.*?)\n\[', conf)
            except:
                print("当前配置应用触发动作未填写")
                break

            HOST = re.findall('HOST=(.*?)\n', conf)  # HOST
            if HOST[0].split('|'):
                host = HOST[0].split('|')
                hostCount = hostCount + (len(host) - 1)
            URL = re.findall('URL=(.*?)\n', conf)  # URL
            if URL[0].split('|'):
                url = URL[0].split('|')
                urlCount = urlCount + (len(url) - 1)
            runVersion = re.findall('RUN_VERSION=(.*?)\n', conf)  # PR版本

            T0_field = re.findall('T_\(0\)\=NONE\|STRING\((.*?)\)\n', conf)[0].split(',')  # 必填字段预处理
            # for T0_field_remove in T0_field:
            #     if T0_field_remove.find('30') != -1:
            #         T0_field.remove(T0_field_remove)  # 去除30*字段
            #     if T0_field_remove.find('40') != -1:
            #         T0_field.remove(T0_field_remove)  # 去除40*字段
            #     if T0_field_remove.find('100') != -1:
            #         T0_field.remove(T0_field_remove)  # 去除100*字段
            # 校验T0字段
            try:
                if T0_field[0] == '':
                    print(line2 + "当前配置T_(0)字段错误，请修改配置]")
                    break
            except IndexError:
                print(line2 + "当前配置T_(0)字段错误，请修改配置]")
                break
            protocolType = re.findall('T_\(6\)\=NONE\|STRING\((.*?)\)\n', conf)  # 大协议类型
            # 查询ACTION字段下标
            try:
                actionSql = "select 字段下标 from" + "`" + protocolType[0] + "`" + " where 字段英文名 ='ACTION';"
                cursor_field.execute(actionSql)
                action = cursor_field.fetchone()
                action_temp = action["字段下标"]
            except TypeError:
                actionSql = "select 字段下标 from" + "`" + protocolType[0] + "`" + " where 字段英文名 ='APP_ACTION';"
                cursor_field.execute(actionSql)
                action = cursor_field.fetchone()
                action_temp = action["字段下标"]
            # 根据配置id查询用户
            conf_Id = re.findall('T_\(998\)\=NONE\|STRING\((.*?)\)\n', conf)  # 配置id
            userid = (int(conf_Id[0]) & 0x1fffc0) >> 6
            try:
                userSql = "select name from user where id =" + "'" + str(userid) + "';"
                cursor_user.execute(userSql)
                userid_temp = cursor_user.fetchone()
                username = userid_temp["name"]
            except:
                print("当前工号未录入数据库中，请录入")
                break
            confIdTemp.append(conf_Id[0])
            global protocolValue
            protocolValue = re.findall('N_\(38\)\=NONE\|STRING\((.*?)\)\n', conf)  # 小协议值
            actionValue = re.findall('N_\(' + str(action_temp) + '\)\=NONE\|STRING\((.*?)\)\n', conf)  # 动作值
            # 查询应用及应用类别
            try:
                protocolSql = "select 应用类别,应用 from" + "`" + protocolType[0] + "`" + " where 协议编号 =" + "'" + \
                              protocolValue[
                                  0] + "';"
                cursor_protocol.execute(protocolSql)
                protocolResult = cursor_protocol.fetchone()
                protocolTypeName = protocolResult["应用类别"]
                protocolValueName = protocolResult["应用"].replace('/', '')
            except TypeError:
                print("当前配置协议值未录入数据库中，请录入")
                break
            # 查询动作描述
            try:
                actionDescSql = "select 动作描述 from" + "`" + protocolType[0] + "`" + " where 动作编号 =" + "'" + actionValue[
                    0] + "';"
                cursor_action.execute(actionDescSql)
                actionDesc = cursor_action.fetchone()
                actionName = actionDesc["动作描述"]
            except TypeError:
                print("当前配置动作值未录入数据库中，请录入")
                break
            # N字段预处理，去除N38及M字段添加到N字段中
            N_field = re.findall('N_\((.*?)\)', conf)  # 字段
            N_field.remove('38')
            if re.findall('M_\((.*?)\)', conf):  # 多条输出字段
                M_field = re.findall('M_\((.*?)\)', conf)
                for m_field in M_field:
                    N_field.append(m_field)

            # 数据录入
            ws.cell(row=i, column=1).value = '烽火通信科技股份有限公司'
            ws.cell(row=i, column=2).value = protocolType[0]
            ws.cell(row=i, column=3).value = protocolTypeName
            ws.cell(row=i, column=4).value = protocolValue[0]
            ws.cell(row=i, column=5).value = protocolValueName
            ws.cell(row=i, column=7).value = toolType
            ws.cell(row=i, column=8).value = appVersion
            ws.cell(row=i, column=9).value = actionValue[0]
            ws.cell(row=i, column=10).value = actionName
            ws.cell(row=i, column=12).value = actionTrigger[0]
            ws.cell(row=i, column=14).value = "版本更新"
            ws.cell(row=i, column=15).value = "HTTP"
            ws.cell(row=i, column=17).value = date
            ws.cell(row=i, column=18).value = username
            # 判断分析方式
            if re.findall('SPECIAL_FLAG=(.*?)\n', conf):
                ws.cell(row=i, column=19).value = "全文分析"
            else:
                ws.cell(row=i, column=19).value = "应用分析"
            ws.cell(row=i, column=20).value = conf_Id[0]
            ws.cell(row=i, column=21).value = date
            ws.cell(row=i, column=22).value = username
            ws.cell(row=i, column=23).value = "2.0"
            ws.cell(row=i, column=26).value = runVersion[0]
            # 判断还原功能类型
            if re.findall('SPECIAL_FLAG=(.*?)\n', conf):
                ws.cell(row=i, column=27).value = "PK功能"
            else:
                ws.cell(row=i, column=27).value = "标准功能"
            ws.cell(row=i, column=28).value = "手工录入"
            ws.cell(row=i, column=29).value = "已实现未测试"
            ws.cell(row=i, column=30).value = "是"
            ws.cell(row=i, column=31).value = "未逆向"
            ws.cell(row=i, column=39).value = filename
            ws.cell(row=i, column=40).value = "明文"
            fieldResult = []
            # 查询N字段中英文及录入

            try:
                j = 52
                k = 53
                m = 54
                for field in N_field:
                    fieldSql = "select 字段英文名,字段中文名 from" + "`" + protocolType[
                        0] + "`" + "  where 字段下标 =" + "'" + field + "';"
                    cursor_field.execute(fieldSql)
                    field_temp = cursor_field.fetchone()
                    fieldChName = field_temp["字段中文名"]
                    if fieldChName in ['动作', '系统类型', '工具类型', '工具名称', '业务系统', '应用名称', '坐标系','操作系统类型','用户代理(备注：这个字段用来填浏览器版本，操作系统类型等等的)','用户代理']:
                        continue
                    j = j + 3
                    k = k + 3
                    m = m + 3
                    fieldResult.append(fieldChName)
                    fieldEnName = field_temp["字段英文名"]
                    if re.findall('N_\(' + field + '\)=.*?\n', conf):
                        N_fieldSearch = re.findall('N_\(' + field + '\)=.*?\n', conf)
                        fieldCount.append(N_fieldSearch[0])
                    if re.findall('M_\(' + field + '\)=.*?\n', conf):
                        M_fieldSearch = re.findall('M_\(' + field + '\)=.*?\n', conf)
                        fieldCount.append(M_fieldSearch[0])
                    ws.cell(row=1, column=j).value = "字段名"
                    ws.cell(row=1, column=j).font = font
                    ws.cell(row=i, column=j).value = fieldEnName
                    ws.cell(row=1, column=k).value = "字段描述"
                    ws.cell(row=1, column=k).font = font
                    ws.cell(row=i, column=k).value = fieldChName
                    ws.cell(row=1, column=m).value = "字段属性"
                    ws.cell(row=1, column=m).font = font
                    # 判断字段属性
                    if field in T0_field:
                        ws.cell(row=i, column=m).value = "必填"
                    else:
                        ws.cell(row=i, column=m).value = "可选"

            except:
                print("当前配置有字段未录入数据库中，请录入")
            # 判断实现方式
            if fieldResult == []:
                ws.cell(row=i, column=16).value = "HTTP识别"
                confDomainCount += 1
            else:
                ws.cell(row=i, column=16).value = "HTTP配置"
            ws.cell(row=i, column=53).value = str(fieldResult).replace('[', '').replace("'", '').replace(']',
                                                                                                         '').replace(
                ',', '、').replace(' ', '')
            if re.findall('T_\(30.\)=.*?\n', conf):
                T_field = re.findall('T_\(30.\)=.*?\n', conf)
                fieldCount.extend(T_field)
        # 校验配置ID
        for key, value in dict(Counter(confIdTemp)).items():
            if value > 1:
                print("当前配置ID重复：", key, "\n重复次数：", value)
        print("当前配置识别字段数：", confDomainCount)
        print("当前配置HOST及URL字段数：", hostCount + urlCount)
        print("当前配置N-M-T字段数：", len(set(fieldCount)))
        print("总计：", confDomainCount + hostCount + urlCount + len(set(fieldCount)))
        cursor_field.close()
        conn_field.close()
        cursor_action.close()
        conn_action.close()
        cursor_protocol.close()
        conn_protocol.close()
        cursor_user.close()
        conn_user.close()
        wb.save('714666111_' + protocolValue[0] + '_' + createDate + ".xlsx")

    except Exception as ex:
        print("出现异常%s" % ex)



num = 0
while num == 0:
    str_file = os.getcwd()
    for root, appnames, files in os.walk(str_file):
        break
    print("请选择文件名：")
    for i in range(len(files)):
        print(i, files[i])
    filenum = int(input())
    filename = files[filenum]
    searchFileContent(filename)
    num = int(input("任意键退出，0继续..."))